"""
File Metadata Extraction System.

This module provides functionality to extract file system metadata and format-specific
metadata from various file types including documents, spreadsheets, and text files.
"""

import os
import re
import stat
import chardet
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Callable, Any, Tuple, Union

from src.core.logging import LoggerMixin
from .exceptions import MetadataExtractionError, FileSystemError
from .parsers import ParserFactory
from .scanner import ScanResult
from .metadata_types import (
    FileMetadata, BasicFileMetadata, FormatSpecificMetadata,
    PDFMetadata, WordMetadata, ExcelMetadata, TextMetadata, ImageMetadata,
    MetadataStats
)


class MetadataExtractor(LoggerMixin):
    """
    File Metadata Extraction System.
    
    Extracts both basic file system metadata and format-specific metadata
    from various file types. Integrates with existing parsers and provides
    batch processing capabilities.
    """
    
    def __init__(self, 
                 parser_factory: Optional[ParserFactory] = None,
                 enable_format_metadata: bool = True,
                 max_file_size: Optional[int] = None,
                 timeout_seconds: float = 30.0):
        """
        Initialize the metadata extractor.
        
        Args:
            parser_factory: Factory for file parsers (optional)
            enable_format_metadata: Whether to extract format-specific metadata
            max_file_size: Maximum file size to process (in bytes)
            timeout_seconds: Timeout for metadata extraction per file
        """
        self.parser_factory = parser_factory or ParserFactory()
        self.enable_format_metadata = enable_format_metadata
        self.max_file_size = max_file_size
        self.timeout_seconds = timeout_seconds
        
        # Statistics
        self.stats = {
            "files_processed": 0,
            "successful_extractions": 0,
            "failed_extractions": 0,
            "format_metadata_extractions": 0,
            "total_extraction_time": 0.0
        }
        
        self.logger.info("Metadata Extractor initialized successfully")
    
    def extract_metadata(self, file_path: Union[str, Path]) -> FileMetadata:
        """
        Extract complete metadata from a file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Complete file metadata
            
        Raises:
            MetadataExtractionError: If metadata extraction fails
        """
        start_time = datetime.now()
        file_path = Path(file_path)
        errors = []
        
        try:
            self.logger.debug(f"Extracting metadata for: {file_path}")
            
            # Check file existence
            if not file_path.exists():
                raise MetadataExtractionError(f"File does not exist: {file_path}")
            
            # Check file size limit
            if self.max_file_size:
                file_size = file_path.stat().st_size
                if file_size > self.max_file_size:
                    raise MetadataExtractionError(
                        f"File too large ({file_size} bytes): {file_path}"
                    )
            
            # Extract basic metadata
            basic_metadata = self._extract_basic_metadata(file_path)
            
            # Extract format-specific metadata
            format_metadata = None
            if self.enable_format_metadata:
                try:
                    format_metadata = self._extract_format_metadata(file_path)
                    if format_metadata:
                        self.stats["format_metadata_extractions"] += 1
                except Exception as e:
                    error_msg = f"Format metadata extraction failed: {e}"
                    errors.append(error_msg)
                    self.logger.warning(f"{error_msg} for {file_path}")
            
            # Create metadata object
            metadata = FileMetadata(
                path=file_path,
                basic_metadata=basic_metadata,
                format_metadata=format_metadata,
                extraction_timestamp=datetime.now(),
                extraction_errors=errors
            )
            
            # Update statistics
            self.stats["files_processed"] += 1
            self.stats["successful_extractions"] += 1
            
            end_time = datetime.now()
            extraction_time = (end_time - start_time).total_seconds()
            self.stats["total_extraction_time"] += extraction_time
            
            self.logger.debug(f"Metadata extraction completed for {file_path} in {extraction_time:.3f}s")
            
            return metadata
            
        except Exception as e:
            self.stats["files_processed"] += 1
            self.stats["failed_extractions"] += 1
            
            self.logger.error(f"Metadata extraction failed for {file_path}: {e}")
            raise MetadataExtractionError(f"Failed to extract metadata from {file_path}: {e}")
    
    def _extract_basic_metadata(self, file_path: Path) -> BasicFileMetadata:
        """Extract basic file system metadata."""
        try:
            # Get file stats
            stat_result = file_path.stat()
            
            # File size
            size = stat_result.st_size
            
            # Time information
            modification_time = datetime.fromtimestamp(stat_result.st_mtime)
            access_time = datetime.fromtimestamp(stat_result.st_atime)
            
            # Creation time (platform-specific)
            try:
                # Try to get birth time (creation time)
                if hasattr(stat_result, 'st_birthtime'):
                    creation_time = datetime.fromtimestamp(stat_result.st_birthtime)
                else:
                    # Fallback to ctime
                    creation_time = datetime.fromtimestamp(stat_result.st_ctime)
            except (AttributeError, OSError):
                creation_time = modification_time
            
            # File permissions
            permissions = stat.filemode(stat_result.st_mode)
            
            # File type
            file_type = self._determine_file_type(stat_result.st_mode)
            
            # Ownership information
            owner_uid = stat_result.st_uid
            group_gid = stat_result.st_gid
            
            # Additional metadata
            is_symlink = file_path.is_symlink()
            inode = getattr(stat_result, 'st_ino', None)
            device_id = getattr(stat_result, 'st_dev', None)
            hard_link_count = getattr(stat_result, 'st_nlink', None)
            
            return BasicFileMetadata(
                size=size,
                creation_time=creation_time,
                modification_time=modification_time,
                access_time=access_time,
                permissions=permissions,
                file_type=file_type,
                owner_uid=owner_uid,
                group_gid=group_gid,
                is_symlink=is_symlink,
                inode=inode,
                device_id=device_id,
                hard_link_count=hard_link_count
            )
            
        except Exception as e:
            raise MetadataExtractionError(f"Failed to extract basic metadata: {e}")
    
    def _determine_file_type(self, mode: int) -> str:
        """Determine file type from stat mode."""
        if stat.S_ISREG(mode):
            return "regular"
        elif stat.S_ISDIR(mode):
            return "directory"
        elif stat.S_ISLNK(mode):
            return "symlink"
        elif stat.S_ISCHR(mode):
            return "character_device"
        elif stat.S_ISBLK(mode):
            return "block_device"
        elif stat.S_ISFIFO(mode):
            return "fifo"
        elif stat.S_ISSOCK(mode):
            return "socket"
        else:
            return "unknown"
    
    def _extract_format_metadata(self, file_path: Path) -> Optional[FormatSpecificMetadata]:
        """Extract format-specific metadata based on file type."""
        if not self.parser_factory.is_supported(file_path):
            return None
        
        try:
            extension = file_path.suffix.lower()
            
            if extension == '.pdf':
                return self._extract_pdf_metadata(file_path)
            elif extension in ['.docx', '.doc']:
                return self._extract_word_metadata(file_path)
            elif extension in ['.xlsx', '.xls']:
                return self._extract_excel_metadata(file_path)
            elif extension in ['.txt', '.md', '.rst', '.py', '.js', '.html', '.css', '.json', '.xml', '.yaml', '.yml']:
                return self._extract_text_metadata(file_path)
            elif extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.svg']:
                return self._extract_image_metadata(file_path)
            else:
                self.logger.debug(f"No format-specific metadata extractor for: {extension}")
                return None
                
        except Exception as e:
            self.logger.warning(f"Format metadata extraction failed for {file_path}: {e}")
            return None
    
    def _extract_pdf_metadata(self, file_path: Path) -> Optional[PDFMetadata]:
        """Extract metadata from PDF files."""
        try:
            # Use parser factory to get PDF parser
            parser = self.parser_factory.get_parser(file_path)
            
            # Basic extraction to check if file is readable
            content = parser.extract_text(file_path)
            
            # Try to extract detailed metadata
            page_count = 0
            title = None
            author = None
            creator = None
            producer = None
            creation_date = None
            modification_date = None
            encrypted = False
            version = None
            
            try:
                # Try to import PyPDF2 for detailed metadata
                import PyPDF2
                
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    
                    page_count = len(pdf_reader.pages)
                    encrypted = pdf_reader.is_encrypted
                    
                    if hasattr(pdf_reader, 'metadata') and pdf_reader.metadata:
                        metadata = pdf_reader.metadata
                        title = metadata.get('/Title')
                        author = metadata.get('/Author')
                        creator = metadata.get('/Creator')
                        producer = metadata.get('/Producer')
                        
                        # Parse dates
                        if metadata.get('/CreationDate'):
                            try:
                                creation_date = self._parse_pdf_date(metadata['/CreationDate'])
                            except:
                                pass
                        
                        if metadata.get('/ModDate'):
                            try:
                                modification_date = self._parse_pdf_date(metadata['/ModDate'])
                            except:
                                pass
                    
                    # Try to get PDF version
                    if hasattr(pdf_reader, 'pdf_header'):
                        version = pdf_reader.pdf_header
                        
            except ImportError:
                self.logger.debug("PyPDF2 not available, using basic PDF metadata")
                # Estimate page count from content length (very rough)
                if content:
                    page_count = max(1, len(content) // 2000)
            except Exception as e:
                self.logger.debug(f"Failed to extract detailed PDF metadata: {e}")
                page_count = 1  # Fallback
            
            return PDFMetadata(
                page_count=page_count,
                title=title,
                author=author,
                creator=creator,
                producer=producer,
                creation_date=creation_date,
                modification_date=modification_date,
                encrypted=encrypted,
                version=version
            )
            
        except Exception as e:
            self.logger.warning(f"PDF metadata extraction failed: {e}")
            return None
    
    def _extract_word_metadata(self, file_path: Path) -> Optional[WordMetadata]:
        """Extract metadata from Word documents."""
        try:
            # Use parser factory to get Word parser
            parser = self.parser_factory.get_parser(file_path)
            content = parser.extract_text(file_path)
            
            # Basic text analysis
            word_count = len(content.split()) if content else 0
            character_count = len(content) if content else 0
            paragraph_count = len([p for p in content.split('\n\n') if p.strip()]) if content else 0
            
            # Try to extract detailed metadata using python-docx
            title = None
            author = None
            last_modified_by = None
            creation_date = None
            last_modified = None
            page_count = 0
            
            try:
                import docx
                
                if file_path.suffix.lower() == '.docx':
                    doc = docx.Document(file_path)
                    
                    # Core properties
                    core_props = doc.core_properties
                    title = core_props.title
                    author = core_props.author
                    last_modified_by = core_props.last_modified_by
                    creation_date = core_props.created
                    last_modified = core_props.modified
                    
                    # Estimate page count (rough)
                    page_count = max(1, len(doc.paragraphs) // 30)
                    
            except ImportError:
                self.logger.debug("python-docx not available, using basic Word metadata")
            except Exception as e:
                self.logger.debug(f"Failed to extract detailed Word metadata: {e}")
            
            return WordMetadata(
                word_count=word_count,
                page_count=page_count,
                character_count=character_count,
                character_count_with_spaces=character_count,
                paragraph_count=paragraph_count,
                title=title,
                author=author,
                last_modified_by=last_modified_by,
                creation_date=creation_date,
                last_modified=last_modified
            )
            
        except Exception as e:
            self.logger.warning(f"Word metadata extraction failed: {e}")
            return None
    
    def _extract_excel_metadata(self, file_path: Path) -> Optional[ExcelMetadata]:
        """Extract metadata from Excel files."""
        try:
            # Use parser factory to get Excel parser
            parser = self.parser_factory.get_parser(file_path)
            content = parser.extract_text(file_path)
            
            sheet_names = []
            sheet_count = 0
            total_rows = 0
            total_columns = 0
            title = None
            author = None
            creation_date = None
            last_modified = None
            
            try:
                import openpyxl
                
                if file_path.suffix.lower() == '.xlsx':
                    workbook = openpyxl.load_workbook(file_path, data_only=True)
                    
                    sheet_names = workbook.sheetnames
                    sheet_count = len(sheet_names)
                    
                    # Count total rows and columns
                    for sheet in workbook.worksheets:
                        if sheet.max_row:
                            total_rows += sheet.max_row
                        if sheet.max_column:
                            total_columns += sheet.max_column
                    
                    # Core properties
                    props = workbook.properties
                    title = props.title
                    author = props.creator
                    creation_date = props.created
                    last_modified = props.modified
                    
            except ImportError:
                self.logger.debug("openpyxl not available, using basic Excel metadata")
                # Try to estimate from content
                if content:
                    lines = content.split('\n')
                    sheet_count = 1
                    total_rows = len(lines)
            except Exception as e:
                self.logger.debug(f"Failed to extract detailed Excel metadata: {e}")
            
            return ExcelMetadata(
                sheet_names=sheet_names,
                sheet_count=sheet_count,
                total_rows=total_rows,
                total_columns=total_columns,
                title=title,
                author=author,
                creation_date=creation_date,
                last_modified=last_modified
            )
            
        except Exception as e:
            self.logger.warning(f"Excel metadata extraction failed: {e}")
            return None
    
    def _extract_text_metadata(self, file_path: Path) -> Optional[TextMetadata]:
        """Extract metadata from text files."""
        try:
            # Detect encoding
            encoding = self._detect_encoding(file_path)
            
            # Read file content
            with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                content = f.read()
            
            # Text analysis
            line_count = len(content.splitlines())
            word_count = len(content.split())
            character_count = len(content)
            
            # Check for BOM
            has_bom = False
            try:
                with open(file_path, 'rb') as f:
                    first_bytes = f.read(4)
                    if first_bytes.startswith(b'\xff\xfe') or first_bytes.startswith(b'\xfe\xff') or first_bytes.startswith(b'\xef\xbb\xbf'):
                        has_bom = True
            except:
                pass
            
            # Detect line ending type
            line_ending_type = self._detect_line_endings(content)
            
            return TextMetadata(
                line_count=line_count,
                word_count=word_count,
                character_count=character_count,
                encoding=encoding,
                has_bom=has_bom,
                line_ending_type=line_ending_type
            )
            
        except Exception as e:
            self.logger.warning(f"Text metadata extraction failed: {e}")
            return None
    
    def _extract_image_metadata(self, file_path: Path) -> Optional[ImageMetadata]:
        """Extract metadata from image files."""
        try:
            # Try to use PIL/Pillow for image metadata
            try:
                from PIL import Image
                
                with Image.open(file_path) as img:
                    width, height = img.size
                    format_name = img.format
                    color_mode = img.mode
                    
                    # Check for transparency
                    has_transparency = (
                        'transparency' in img.info or
                        color_mode in ('RGBA', 'LA') or
                        (color_mode == 'P' and 'transparency' in img.info)
                    )
                    
                    # Get DPI if available
                    dpi = None
                    if hasattr(img, 'info') and 'dpi' in img.info:
                        dpi = img.info['dpi']
                    
                    return ImageMetadata(
                        width=width,
                        height=height,
                        format=format_name,
                        color_mode=color_mode,
                        has_transparency=has_transparency,
                        dpi=dpi
                    )
                    
            except ImportError:
                self.logger.debug("PIL/Pillow not available for image metadata")
                return None
            
        except Exception as e:
            self.logger.warning(f"Image metadata extraction failed: {e}")
            return None
    
    def _detect_encoding(self, file_path: Path) -> str:
        """Detect file encoding."""
        try:
            with open(file_path, 'rb') as f:
                raw_data = f.read(8192)  # Read first 8KB
                result = chardet.detect(raw_data)
                return result['encoding'] or 'utf-8'
        except:
            return 'utf-8'
    
    def _detect_line_endings(self, content: str) -> str:
        """Detect line ending type in text content."""
        if not content:
            return "unknown"
        
        crlf_count = content.count('\r\n')
        lf_count = content.count('\n') - crlf_count
        cr_count = content.count('\r') - crlf_count
        
        if crlf_count > 0 and lf_count == 0 and cr_count == 0:
            return "windows"
        elif lf_count > 0 and crlf_count == 0 and cr_count == 0:
            return "unix"
        elif cr_count > 0 and crlf_count == 0 and lf_count == 0:
            return "mac"
        elif crlf_count > 0 or lf_count > 0 or cr_count > 0:
            return "mixed"
        else:
            return "unknown"
    
    def _parse_pdf_date(self, date_str: str) -> datetime:
        """Parse PDF date string to datetime."""
        # PDF date format: D:YYYYMMDDHHmmSSOHH'mm
        if date_str.startswith('D:'):
            date_str = date_str[2:]
        
        # Extract basic date/time
        if len(date_str) >= 14:
            year = int(date_str[0:4])
            month = int(date_str[4:6])
            day = int(date_str[6:8])
            hour = int(date_str[8:10])
            minute = int(date_str[10:12])
            second = int(date_str[12:14])
            
            return datetime(year, month, day, hour, minute, second)
        
        return datetime.now()
    
    def extract_metadata_batch(self, 
                              scan_result: ScanResult,
                              progress_callback: Optional[Callable[[int, int, Path], None]] = None) -> List[FileMetadata]:
        """
        Extract metadata from multiple files.
        
        Args:
            scan_result: Scan result containing files to process
            progress_callback: Optional callback for progress updates
            
        Returns:
            List of extracted metadata
        """
        results = []
        errors = []
        
        self.logger.info(f"Starting batch metadata extraction for {len(scan_result.files)} files")
        
        for i, file_path in enumerate(scan_result.files):
            try:
                metadata = self.extract_metadata(file_path)
                results.append(metadata)
                
                if progress_callback:
                    progress_callback(i + 1, len(scan_result.files), file_path)
                    
            except Exception as e:
                error_entry = {
                    "path": str(file_path),
                    "error": str(e),
                    "type": "metadata_extraction"
                }
                errors.append(error_entry)
                self.logger.error(f"Failed to extract metadata from {file_path}: {e}")
        
        # Add errors to scan result
        scan_result.errors.extend(errors)
        
        self.logger.info(f"Batch metadata extraction completed: {len(results)} successful, {len(errors)} failed")
        
        return results
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get extraction statistics."""
        stats = self.stats.copy()
        
        if stats["files_processed"] > 0:
            stats["success_rate"] = stats["successful_extractions"] / stats["files_processed"]
            stats["average_extraction_time"] = stats["total_extraction_time"] / stats["files_processed"]
        else:
            stats["success_rate"] = 0.0
            stats["average_extraction_time"] = 0.0
        
        return stats
    
    def reset_statistics(self) -> None:
        """Reset extraction statistics."""
        self.stats = {
            "files_processed": 0,
            "successful_extractions": 0,
            "failed_extractions": 0,
            "format_metadata_extractions": 0,
            "total_extraction_time": 0.0
        }
    
    def generate_metadata_stats(self, metadata_list: List[FileMetadata]) -> MetadataStats:
        """Generate statistics from extracted metadata."""
        if not metadata_list:
            return MetadataStats(
                total_files=0,
                successful_extractions=0,
                failed_extractions=0,
                total_size=0
            )
        
        total_files = len(metadata_list)
        successful_extractions = len([m for m in metadata_list if not m.extraction_errors])
        failed_extractions = total_files - successful_extractions
        
        total_size = sum(m.basic_metadata.size for m in metadata_list)
        average_file_size = total_size / total_files if total_files > 0 else 0.0
        
        # File type distribution
        file_type_dist = {}
        format_type_dist = {}
        
        largest_file = None
        smallest_file = None
        largest_size = 0
        smallest_size = float('inf')
        
        for metadata in metadata_list:
            # File type distribution
            file_type = metadata.basic_metadata.file_type
            file_type_dist[file_type] = file_type_dist.get(file_type, 0) + 1
            
            # Format type distribution
            if metadata.format_metadata:
                format_type = metadata.format_metadata.to_dict().get("type", "unknown")
                format_type_dist[format_type] = format_type_dist.get(format_type, 0) + 1
            
            # Size tracking
            size = metadata.basic_metadata.size
            if size > largest_size:
                largest_size = size
                largest_file = metadata.path
            if size < smallest_size:
                smallest_size = size
                smallest_file = metadata.path
        
        return MetadataStats(
            total_files=total_files,
            successful_extractions=successful_extractions,
            failed_extractions=failed_extractions,
            total_size=total_size,
            file_type_distribution=file_type_dist,
            format_type_distribution=format_type_dist,
            average_file_size=average_file_size,
            largest_file=largest_file,
            smallest_file=smallest_file
        )